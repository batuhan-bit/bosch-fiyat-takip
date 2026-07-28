from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import ReferenceValue
from .parsing import looks_like_model, normalize_model


def _period_from_filename(path: Path) -> str:
    match = re.search(r"(Ocak|Şubat|Mart|Nisan|Mayıs|Haziran|Temmuz|Ağustos|Eylül|Ekim|Kasım|Aralık)\s+(20\d{2})", path.stem, re.IGNORECASE)
    return f"{match.group(1).title()} {match.group(2)}" if match else ""


def load_wholesale(path: Path) -> dict[str, ReferenceValue]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, ReferenceValue] = {}
    period = _period_from_filename(path)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if len(row) < 3 or not looks_like_model(row[1]) or not isinstance(row[2], (int, float)):
                continue
            model = normalize_model(row[1])
            result[model] = ReferenceValue(model, float(row[2]), "Toptan", period, path.name)
    return result


def _support_type(sheet_name: str) -> str | None:
    normalized = normalize_model(sheet_name)
    if "AILE" in normalized or "BAKAN" in normalized:
        return None
    if "FIRSAT" in normalized or "PHASE" in normalized:
        return "Phase-out"
    return "BSP"


def load_support(path: Path) -> dict[str, ReferenceValue]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, ReferenceValue] = {}
    period = _period_from_filename(path)
    for sheet in workbook.worksheets:
        support_type = _support_type(sheet.title)
        if not support_type:
            continue
        for row in sheet.iter_rows(values_only=True):
            if len(row) < 4 or not looks_like_model(row[2]) or not isinstance(row[3], (int, float)):
                continue
            model = normalize_model(row[2])
            if model in result:
                raise ValueError(f"{model} hem {result[model].source_type} hem {support_type} listesinde bulunuyor.")
            result[model] = ReferenceValue(model, float(row[3]), support_type, period, path.name)
    return result

